"""Read-only structural audit for the Manila student workbook archive."""

from __future__ import annotations

import json
import re
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


SUMMARY_FILES = {
    "college ministry master file.xlsx",
    "with honor students.xlsx",
}


def clean(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip()


def normalized_label(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean(value).casefold())


def first_value_to_right(sheet, row: int, start_column: int = 2) -> str:
    for column in range(start_column, min(sheet.max_column, 16) + 1):
        value = clean(sheet.cell(row, column).value)
        if value:
            return value
    return ""


def inspect_workbook(path: Path) -> dict:
    media = []
    with zipfile.ZipFile(path) as archive:
        media = [
            name
            for name in archive.namelist()
            if name.casefold().startswith("xl/media/")
            and not name.endswith("/")
        ]

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = max(
            workbook.worksheets,
            key=lambda candidate: candidate.max_row * candidate.max_column,
        )
        fields = {}
        transaction_header = None
        for row in range(1, min(sheet.max_row, 80) + 1):
            label = normalized_label(sheet.cell(row, 1).value)
            if label:
                fields.setdefault(label, first_value_to_right(sheet, row))
            row_labels = {
                normalized_label(sheet.cell(row, column).value)
                for column in range(1, min(sheet.max_column, 16) + 1)
            }
            if "date" in row_labels and (
                "particulars" in row_labels
                or "debit" in row_labels
                or "credit" in row_labels
            ):
                transaction_header = row

        transaction_rows = []
        if transaction_header:
            for row in range(transaction_header + 1, sheet.max_row + 1):
                values = [
                    sheet.cell(row, column).value
                    for column in range(1, min(sheet.max_column, 12) + 1)
                ]
                populated = [value for value in values if clean(value)]
                if len(populated) >= 2:
                    transaction_rows.append(
                        [clean(value) for value in values]
                    )

        return {
            "file": path.name,
            "sheet": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "fields": fields,
            "media": media,
            "transaction_header": transaction_header,
            "transaction_rows": transaction_rows,
        }
    finally:
        workbook.close()


def main() -> int:
    root = Path(sys.argv[1]).resolve()
    files = sorted(
        path
        for path in root.glob("*.xlsx")
        if path.name.casefold() not in SUMMARY_FILES
    )
    records = [inspect_workbook(path) for path in files]

    by_name = defaultdict(list)
    for record in records:
        name = (
            record["fields"].get("studentsname")
            or record["fields"].get("studentname")
            or ""
        )
        by_name[normalized_label(name)].append(record["file"])

    report = {
        "workbooks": len(records),
        "with_profile_name": sum(
            bool(
                record["fields"].get("studentsname")
                or record["fields"].get("studentname")
            )
            for record in records
        ),
        "with_media": sum(bool(record["media"]) for record in records),
        "media_entries": sum(len(record["media"]) for record in records),
        "with_transaction_header": sum(
            record["transaction_header"] is not None for record in records
        ),
        "with_transaction_rows": sum(
            bool(record["transaction_rows"]) for record in records
        ),
        "transaction_rows": sum(
            len(record["transaction_rows"]) for record in records
        ),
        "duplicate_profile_names": {
            name: files
            for name, files in sorted(by_name.items())
            if name and len(files) > 1
        },
        "media_extensions": Counter(
            Path(media).suffix.casefold()
            for record in records
            for media in record["media"]
        ),
        "samples_with_media": [
            {
                "file": record["file"],
                "media": record["media"],
            }
            for record in records
            if record["media"]
        ][:20],
        "samples_with_transactions": [
            {
                "file": record["file"],
                "rows": record["transaction_rows"][:4],
            }
            for record in records
            if record["transaction_rows"]
        ][:20],
        "missing_profile_names": [
            record["file"]
            for record in records
            if not (
                record["fields"].get("studentsname")
                or record["fields"].get("studentname")
            )
        ],
    }
    print(json.dumps(report, indent=2, default=dict))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
